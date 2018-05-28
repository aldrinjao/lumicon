import os
from openpyxl import load_workbook
import pprint


# proj list
projidlist = set([])
empidlist = set([])

# dictionary of projects by ID
byproj = {}
byprojname = {}

# byemployee
byemp = {}
byempname = {}

# records

records = {}





def setemptyWBid(empid,projectid):
	records.update({empid:{}})	
	records[empid].update({projectid:[0] * 109})	

def setWBid(empid,projectid):
	records[empid].update({projectid:[0] * 109})	


def openWorkbook(file):

	wb = load_workbook(filename=file, read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		
		ws = wb[sheet]
		projectID = int(ws['AG1'].value)
		projectname = str(ws['R1'].value)

		empID = str(ws['R3'].value)
		empName = str(ws['R2'].value)

		if (projectID not in projidlist):
			projidlist.add(projectID)
			setWBidPerProject(projectID,projectname)

		# read through the file then add the row

		for x in xrange(6,109):
			cellAddress = "AK" + str(x)			
			byproj[projectID][x] += float(ws[cellAddress].value)

		# 


		if (empID not in empidlist):
			empidlist.add(empID)
			setWBidPerEmp(empID,empName)


		for x in xrange(6,109):
			cellAddress = "AK" + str(x)			
			byemp[empID][x] += float(ws[cellAddress].value)

#################################


def saveWorkbook(file):

	wb = load_workbook(filename=file, read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		
		ws = wb[sheet]
		
		projectID = int(ws['AG1'].value)
		projectname = str(ws['R1'].value)

		empID = str(ws['R3'].value)
		empName = str(ws['R2'].value)


		# initialize the container
		if (empID not in empidlist):
			empidlist.add(empID)
			setemptyWBid(empID,projectID)

		projidlist.add(projectID)

		setWBid(empID,projectID)


		# read through the file then add the row

		for x in xrange(6,109):
			cellAddress = "AK" + str(x)			
			records[empID][projectID][x] = float(ws[cellAddress].value)

		# 
	

#################################


baseDir = "timesheets"
# get inside the directory
os.chdir("timesheets")

# List all files and directories in current directory
folders = os.listdir('.')



for folder in folders:
	os.chdir(folder)
	files =  os.listdir('.')


	for file in files:
		saveWorkbook(file) 

	os.chdir('..')


pp = pprint.PrettyPrinter(indent=4)

for x in records:
	pp.pprint(records)



# # design and engineering

# conceptdesign = {}
# revConceptDesign = {}
# shopDrawings= {}
# reviShopDrawings= {}
# fabDrawings= {}
# revfabDrawings= {}
# mockupDrawings= {}
# revmockupDrawings= {}
# installationDrawing= {}
# revInstallationDrawings= {}
# asBuiltDrawings= {}
# revasBuiltDrawings= {}
# calc= {}
# revCalc= {}
# takeOff= {}
# revtakeOff= {}
# oB= {}
# Trainings= {}
# projectMeeting= {}
# coordination= {}
# siteVisit= {}
# subcontractors= {}
# otherProjectRelatedWorks= {}
# nonChargeableMisc= {}
# onleave= {}


# # consultancy

# predesign ={}
# designRev ={}
# designDev ={}
# tender ={}
# pMU ={}
# constructionDoc ={}
# siteQAQC ={}


# # summarize by project
# for x in projidlist:

# 	conceptdesign.update({x: sum(byproj[x][38:41])})
# 	revConceptDesign.update({x: sum(byproj[x][41:44])})
# 	shopDrawings.update({x: sum(byproj[x][44:51])})
# 	reviShopDrawings.update({x: sum(byproj[x][51:58])})
# 	fabDrawings.update({x: sum(byproj[x][58:63])})
# 	revfabDrawings.update({x: sum(byproj[x][63:68])})
# 	mockupDrawings.update({x: sum(byproj[x][68:71])})
# 	revmockupDrawings.update({x: sum(byproj[x][71:74])})
# 	installationDrawing.update({x: sum(byproj[x][74:77])})
# 	revInstallationDrawings.update({x: sum(byproj[x][77:80])})
# 	asBuiltDrawings.update({x: sum(byproj[x][80:83])})
# 	revasBuiltDrawings.update({x: sum(byproj[x][83:86])})
# 	calc.update({x: sum(byproj[x][86:88])})
# 	revCalc.update({x: sum(byproj[x][88:90])})
# 	takeOff.update({x: sum(byproj[x][90:94])})
# 	revtakeOff.update({x: sum(byproj[x][94:98])})
# 	oB.update({x: byproj[x][98]})
# 	Trainings.update({x: byproj[x][99]})
# 	projectMeeting.update({x: byproj[x][100]})
# 	coordination.update({x: byproj[x][101]})
# 	siteVisit.update({x: byproj[x][102]})
# 	subcontractors.update({x: sum(byproj[x][38:41])})
# 	otherProjectRelatedWorks.update({x: byproj[x][104]})
# 	nonChargeableMisc.update({x: sum(byproj[x][105:108])})
# 	onleave.update({x: byproj[x][108]})


# 	predesign.update({x: byproj[x][6]+byproj[x][13]})
# 	designRev.update({x: byproj[x][7]+byproj[x][14]+byproj[x][20]+byproj[x][23]+byproj[x][26]+byproj[x][29]})
# 	designDev.update({x: byproj[x][8]+byproj[x][15]+byproj[x][21]+byproj[x][24]+byproj[x][27]+byproj[x][30]})
# 	tender.update({x: byproj[x][9]+byproj[x][16]+byproj[x][22]+byproj[x][25]+byproj[x][28]+byproj[x][31]+byproj[x][32]+byproj[x][35]})
# 	pMU.update({x: byproj[x][10]+byproj[x][17]+byproj[x][33]+byproj[x][36]})
# 	constructionDoc.update({x: byproj[x][11]+byproj[x][18]+byproj[x][34]+byproj[x][37]})
# 	siteQAQC.update({x: byproj[x][12]+byproj[x][19]})

# # 
# #  maybe write to xlsx here
# # 
# # 

# print conceptdesign 
# print revConceptDesign 
# print shopDrawings
# print reviShopDrawings
# print fabDrawings
# print revfabDrawings
# print mockupDrawings
# print revmockupDrawings
# print installationDrawing
# print revInstallationDrawings
# print asBuiltDrawings
# print revasBuiltDrawings
# print calc
# print revCalc
# print takeOff
# print revtakeOff
# print oB
# print Trainings
# print projectMeeting
# print coordination
# print siteVisit
# print subcontractors
# print otherProjectRelatedWorks
# print nonChargeableMisc
# print onleave

# print "consultancy"
# print predesign
# print designRev
# print designDev
# print tender
# print pMU
# print constructionDoc
# print siteQAQC

# print "total"
# print sum(nonChargeableMisc.values())
# print sum(onleave.values())


# print byemp

os.chdir('..')
wb = load_workbook(filename='./out.xlsx')
ws = wb.worksheets[0]

rowVar = 3
columnVar = 1

for x in records:
	ws.cell(row=1, column=columnVar).value = x

	for y in records[x]:
		ws.cell(row=2, column=columnVar).value = y

		for z in xrange(6,109):
			ws.cell(row=rowVar, column=columnVar).value = records[x][y][z]
			
			rowVar+=1		
		
		rowVar = 3
		
		columnVar+=1

wb.save('out.xlsx')	

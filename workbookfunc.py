import os
from openpyxl import load_workbook


class Activity:
    name = ""
    code = ""
    category = ""
    billable = ""


def saveWorkbookbyProj(file):

	wb = load_workbook(filename=file, read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		
		ws = wb[sheet]
		
		projectID = int(ws['AH1'].value)

		posID = str(ws['AH3'].value)




		if projectID in recordsbyproj:
			if posID in recordsbyproj[projectID]:

				print "both existing"+ str(projectID) + " " +str(posID)
				# PROJECT and position existing
				for x in xrange(6,rowcount+6):

					activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
					cellAddress = "AL" + str(x)			
					
					recordsbyproj[projectID][posID][activitycode] +=  ws[cellAddress].value
	
			else:

				print "proj existing"+ str(projectID) + " " +str(posID)
				recordsbyproj[projectID][posID]= {}

				# PROJECT existing BUT NO Position
				for x in xrange(6,rowcount+6):
					activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
					cellAddress = "AL" + str(x)				
					
					recordsbyproj[projectID][posID][activitycode]={}				

					recordsbyproj[projectID][posID][activitycode] =  float(ws[cellAddress].value)


		else:
			print "both new"+ str(projectID) + " " +str(posID)

			recordsbyproj[projectID] = {}
			recordsbyproj[projectID][posID]= {}


			# New project and position
			for x in xrange(6,rowcount+6):
				
				activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
				cellAddress = "AL" + str(x)			

				recordsbyproj[projectID][posID][activitycode]={}				

				recordsbyproj[projectID][posID][activitycode] =  float(ws[cellAddress].value)




def saveWorkbookbyPos(file):

	wb = load_workbook(filename=file, read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		
		ws = wb[sheet]
		
		projectID = int(ws['AH1'].value)
		posID = str(ws['AH3'].value)




		if posID in recordsbypos:
			if projectID in recordsbypos[posID]:

				print "both existing"+ str(posID) + " " +str(projectID)
				# PROJECT and position existing
				for x in xrange(6,rowcount+6):

					activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
					cellAddress = "AL" + str(x)			
					
					recordsbypos[posID][projectID][activitycode] +=  ws[cellAddress].value
	
			else:

				print "proj existing"+ str(posID) + " " +str(projectID)
				recordsbypos[posID][projectID]= {}

				# PROJECT existing BUT NO Position
				for x in xrange(6,rowcount+6):
					activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
					cellAddress = "AL" + str(x)				
					
					recordsbypos[posID][projectID][activitycode]={}				

					recordsbypos[posID][projectID][activitycode] =  float(ws[cellAddress].value)


		else:
			print "both new"+ str(posID) + " " +str(projectID)

			recordsbypos[posID] = {}
			recordsbypos[posID][projectID]= {}


			# New project and position
			for x in xrange(6,rowcount+6):
				
				activitycode = str(ws["D" + str(x)].value).encode('ascii', 'ignore')	
				cellAddress = "AL" + str(x)			

				recordsbypos[posID][projectID][activitycode]={}				

				recordsbypos[posID][projectID][activitycode] =  float(ws[cellAddress].value)



#add error handlers


def saveReference():
	global rowcount

	wb = load_workbook(filename="./reference.xlsx", read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		ws = wb[sheet]
		rownum = 3

		if sheet == "activity":
			for row in ws.iter_rows(row_offset=2):
				
				cellAddress = "F" + str(rownum)
				code = ws[cellAddress].value

				temp = Activity()

				temp.code = code

				cellAddress = "G" + str(rownum)
				temp.category = ws[cellAddress].value


				cellAddress = "D" + str(rownum)
				temp.name = ws[cellAddress].value

				cellAddress = "H" + str(rownum)
				temp.billable = ws[cellAddress].value


				activities[code] = temp
				activitylist.add(code)
				
				rownum+=1
				# rowcount+=1

		if sheet == "major":
			for row in ws.iter_rows(row_offset=2):
				

				cellAddress = "B" + str(rownum)
				code = ws[cellAddress].value

				cellAddress = "A" + str(rownum)
				name = ws[cellAddress].value

				major[code] = name
				
				rownum+=1

		if sheet == "project":
			for row in ws.iter_rows(row_offset=2):
				

				cellAddress = "B" + str(rownum)
				code = ws[cellAddress].value

				cellAddress = "A" + str(rownum)
				name = ws[cellAddress].value

				project[code] = name
				
				rownum+=1

		if sheet == "position":
			for row in ws.iter_rows(row_offset=2):
				

				cellAddress = "B" + str(rownum)
				code = ws[cellAddress].value

				cellAddress = "A" + str(rownum)
				name = ws[cellAddress].value

				position[code] = name
				
				rownum+=1



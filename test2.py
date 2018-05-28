import os
from openpyxl import *
import pprint
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
    BarChart
)

from copy import deepcopy


# proj list

projidlist = set([])
posidlist = set([])
activitylist = set([])

# records

recordsbypos = {}
recordsbyproj = {}


#references

position = {}
project = {}
major = {}
majorbypos = {}
majorbyproj = {}
label = {}

year = 2018
month = ""

activities = {}




# define the activity class
class Activity:
    name = ""
    code = ""
    category = ""
    billable = True



def assignKeys(order):

	keyPair=[0]* 2

	keyPair[0] = "AH1"
	keyPair[1] = "AH3"

	if order == 'project':
		keyPair[0] = "AH1"
		keyPair[1] = "AH3"


	if order == 'position':
		keyPair[0] = "AH3"
		keyPair[1] = "AH1"



	return keyPair





def saveWorkbook(file,storage,choice,order):

	wb = load_workbook(filename=file, read_only=True,data_only=True)
	sheets = wb.sheetnames


	for sheet in sheets:
		
		ws = wb[sheet]
		
		keys = assignKeys(order)
		firstkey = str(ws[keys[0]].value)
		secondkey = str(ws[keys[1]].value)

# s4 and ah4
		monthInput = str(ws['S4'].value)
		yearInput = str(ws['AH4'].value)
		
		
		if (monthInput ==  month ) and (yearInput == year):


			if firstkey in storage:
				if secondkey in storage[firstkey]:
	
					# PROJECT and position existing	

					for row in ws.iter_rows(min_row=rowOffset):
						if row[3].value:


							activitycode = str(row[3].value).encode('ascii', 'ignore')

							if choice == 'major':

								activitycode = activities[activitycode].category
								



							if activitycode in storage[firstkey][secondkey]:
								
								
								storage[firstkey][secondkey][activitycode] +=  float(row[37].value)
							
							else:

								storage[firstkey][secondkey][activitycode] =  float(row[37].value)
							
						else:
							break


					
				else:

					storage[firstkey][secondkey]= {}

					# PROJECT existing BUT NO Position
					for row in ws.iter_rows(min_row=rowOffset):
						if row[3].value:

							activitycode = str(row[3].value).encode('ascii', 'ignore')	
							
							if choice == 'major':

								activitycode = activities[activitycode].category 

								if activitycode in storage[firstkey][secondkey]:

									storage[firstkey][secondkey][activitycode] +=  float(row[37].value)

								else:
									storage[firstkey][secondkey][activitycode]={}
									storage[firstkey][secondkey][activitycode] =  float(row[37].value)


							else:

								storage[firstkey][secondkey][activitycode]={}				

								storage[firstkey][secondkey][activitycode] =  float(row[37].value)


						else:
							break

			else:
				
				storage[firstkey] = {}
				storage[firstkey][secondkey]= {}


				# New project and position
				for row in ws.iter_rows(min_row=rowOffset):

					if row[3].value:
					

						activitycode = str(row[3].value).encode('ascii', 'ignore')	
			
						if choice=='major':

							activitycode = activities[activitycode].category 

							if activitycode in storage[firstkey][secondkey]:

								storage[firstkey][secondkey][activitycode] +=  float(row[37].value)

							else:
								storage[firstkey][secondkey][activitycode]={}
								storage[firstkey][secondkey][activitycode] =  float(row[37].value)


						else:

							storage[firstkey][secondkey][activitycode]={}				

							storage[firstkey][secondkey][activitycode] =  float(row[37].value)

						

					else:
						break


	return storage



#add error handlers

def saveReference():

	global month
	global year


	wb = load_workbook(filename="./reference.xlsx", read_only=True,data_only=True)
	sheets = wb.sheetnames

	for sheet in sheets:
		ws = wb[sheet]


		if sheet == "activity":
			for row in ws.iter_rows(row_offset=2):
				
				code = str(row[5].value)

				temp = Activity()

				temp.code = code

				temp.category = row[6].value
				temp.name = row[3].value
				temp.billable = row[7].value

				activities[code] = temp
				activitylist.add(code)
				


		if sheet == "major":
			for row in ws.iter_rows(row_offset=2):
				
				code = str(row[1].value)
				name = row[0].value
				major[code] = str(name).lower()


		if sheet == "configuration":
			for row in ws.iter_rows(row_offset=2):
				
				year = str(row[0].value)
				month = str(row[1].value)


		for row in ws.iter_rows(row_offset=2):
			
			code = str(row[1].value)
			name = row[0].value
			label[code] = str(name).lower()
			



def writeOutput(outfile,records):


####################################

	wb = Workbook()
	summaryColumn = 4
	for record in records:
		sheetname = str(label[record])
		wb.create_sheet(sheetname)
		ws = wb[str(sheetname)]
		wsummary = wb.worksheets[0]


		# write labels

		ws.cell(row=1, column=1).value = "code"
		ws.cell(row=1, column=2).value = "name"		

		columnVar = 1
		rowVar = 2

		for x in sorted(activitylist):
			ws.cell(row=rowVar, column=columnVar).value = activities[x].code
			ws.cell(row=rowVar, column=columnVar+1).value = activities[x].name
			ws.cell(row=rowVar, column=columnVar+2).value = label[str(activities[x].category)]
			
			rowVar += 1

		
		columnVar = 4
		rowVar = 2


		for r in records[record]:
			# r is column
			ws.cell(row=1, column=columnVar).value = label[r]


			for x in xrange(0,len(activitylist)):
	
				code = str(ws.cell(row=rowVar+x, column=1).value)
				
				if code in records[record][r]: 

					ws.cell(row=rowVar+x, column=columnVar).value = records[record][r][code] 
					
					if isinstance(wsummary.cell(row=rowVar+x, column=summaryColumn).value,float):

						wsummary.cell(row=rowVar+x, column=summaryColumn).value += float(records[record][r][code]) 
					
					else:

						wsummary.cell(row=rowVar+x, column=summaryColumn).value = float(records[record][r][code]) 				

			
			columnVar += 1
		summaryColumn+=1			



		for col in ws.columns:
		     max_length = 0
		     column = col[0].column # Get the column name
		     for cell in col:
		         try: # Necessary to avoid error on empty cells
		             if len(str(cell.value)) > max_length:
		                 max_length = len(cell.value)
		         except:
		             pass
		     adjusted_width = (max_length) * 1
		     ws.column_dimensions[column].width = adjusted_width	
####################################


	# write values





	ws = wb.worksheets[0]
	ws.title="summary"
	columnVar = 1
	rowVar = 2



		# write labels



	for x in sorted(activitylist):
		ws.cell(row=rowVar, column=columnVar).value = activities[x].code
		ws.cell(row=rowVar, column=columnVar+1).value = activities[x].name
		ws.cell(row=rowVar, column=columnVar+2).value = label[str(activities[x].category)]	


		rowVar +=1
	# make multiple outputs per project and by per employee
	
	columnVar = 4
	rowVar = 1


	for record in records:

		ws.cell(row=rowVar, column=columnVar).value = label[record]
		columnVar+=1


	columnVar = 4
	rowVar = 2



	for col in ws.columns:
	     max_length = 0
	     column = col[0].column # Get the column name
	     for cell in col:
	         try: # Necessary to avoid error on empty cells
	             if len(str(cell.value)) > max_length:
	                 max_length = len(cell.value)
	         except:
	             pass
	     adjusted_width = (max_length) * 1
	     ws.column_dimensions[column].width = adjusted_width



	wb.save(year+"_"+month+"_"+outfile)




	# write summary


#################################





def writeOutput2(outfile,records):


####################################

	wb = Workbook()
	summaryColumn = 3


	for record in records:
		sheetname = str(label[record])
		wb.create_sheet(sheetname)
		ws = wb[str(sheetname)]
		wsummary = wb.worksheets[0]




		# write labels

		ws.cell(row=1, column=1).value = "code"
		ws.cell(row=1, column=2).value = "name"		

		columnVar = 1
		rowVar = 2

		for x in sorted(major):
			ws.cell(row=rowVar, column=columnVar).value = x
			ws.cell(row=rowVar, column=columnVar+1).value = label[x]
			
			rowVar += 1

		
		columnVar = 3
		rowVar = 2


		for r in records[record]:

			# r is column
			ws.cell(row=1, column=columnVar).value = label[r]

			for x in xrange(0,len(major)):
				

				code = str(ws.cell(row=rowVar+x, column=1).value)
				if code in records[record][r]: 

					ws.cell(row=rowVar+x, column=columnVar).value = records[record][r][code] 
					
					if isinstance(wsummary.cell(row=rowVar+x, column=summaryColumn).value,float):
						
						wsummary.cell(row=rowVar+x, column=summaryColumn).value += float(records[record][r][code])

					
					else:
						
						wsummary.cell(row=rowVar+x, column=summaryColumn).value = float(records[record][r][code]) 
					

			
			columnVar += 1

		for col in ws.columns:
		     max_length = 0
		     column = col[0].column # Get the column name
		     for cell in col:
		         try: # Necessary to avoid error on empty cells
		             if len(str(cell.value)) > max_length:
		                 max_length = len(cell.value)
		         except:
		             pass
		     adjusted_width = (max_length) * 1
		     ws.column_dimensions[column].width = adjusted_width


		summaryColumn+=1			






		data = Reference(ws, min_col=3, min_row=1, max_col=columnVar-1, max_row=6)
		titles = Reference(ws, min_col=2, min_row=2, max_row=6)
		chart = BarChart()
		chart.title = label[record]
		chart.add_data(data=data, titles_from_data=True)
		chart.set_categories(titles)
		chart.x_axis.delete = False
		chart.y_axis.delete = False
		ws.add_chart(chart, "A10")





####################################


	# write values



	ws = wb.worksheets[0]
	ws.title="summary"
	columnVar = 1
	rowVar = 2



# write labels

	ws.cell(row=1, column=1).value = "code"
	ws.cell(row=1, column=2).value = "name"		


	for x in sorted(major):
		ws.cell(row=rowVar, column=columnVar).value = x
		ws.cell(row=rowVar, column=columnVar+1).value = label[x]	


		rowVar +=1
	# make multiple outputs per project and by per employee
	
	columnVar = 1
	rowVar = 1


	for record in records:

		ws.cell(row=rowVar, column=columnVar+2).value = label[record]
		columnVar+=1


	for col in ws.columns:
	     max_length = 0
	     column = col[0].column # Get the column name
	     for cell in col:
	         try: # Necessary to avoid error on empty cells
	             if len(str(cell.value)) > max_length:
	                 max_length = len(cell.value)
	         except:
	             pass
	     adjusted_width = (max_length) * 1
	     ws.column_dimensions[column].width = adjusted_width

	print outfile, columnVar	 


	data = Reference(ws, min_col=3, min_row=1, max_col=columnVar+1, max_row=6)
	titles = Reference(ws, min_col=2, min_row=2, max_row=6)
	chart = BarChart3D()
	chart.title = "3D Bar Chart"
	chart.add_data(data=data, titles_from_data=True)
	chart.set_categories(titles)

	ws.add_chart(chart, "A10")

	chart3 = BarChart()
	chart3 = deepcopy(chart)
	chart3.type = "col"
	chart3.style = 10
	chart3.grouping = "stacked"
	chart3.overlap = 100
	chart3.title = 'Stacked Chart'
	chart3.set_categories(titles)

	ws.add_chart(chart3, "A30")

	chart.title.delete = False
	chart.x_axis.delete = False
	chart.y_axis.delete = False

	chart3.title.delete = False
	chart3.x_axis.delete = False
	chart3.y_axis.delete = False


	wb.save(year+"_"+month+"_"+outfile)


	# write summary

			
#################################

#
# Read reference file and store them
# 

rowcount = 0;
rowOffset = 6

year = ""
month = ""

saveReference()


#################################




# get inside the directory
os.chdir("timesheets")

# # List all files and directories in current directory
folders = os.listdir('.')



for folder in folders:
	os.chdir(folder)
	files =  os.listdir('.')


	for file in files:
		recordsbypos = saveWorkbook(file,recordsbypos,'any','position')
		recordsbyproj = saveWorkbook(file,recordsbyproj,'any','project') 
		majorbypos = saveWorkbook(file,majorbypos,'major','position') 
		majorbyproj = saveWorkbook(file,majorbyproj,'major','project') 
	
	os.chdir('..')


pp = pprint.PrettyPrinter(indent=4)

pp.pprint(majorbypos)
pp.pprint(majorbyproj)


# # go back to main directory

os.chdir('..')
writeOutput("out_pos.xlsx",recordsbypos)
writeOutput("out_prj.xlsx",recordsbyproj)
writeOutput2("out_prj_major.xlsx",majorbyproj)
writeOutput2("out_pos_major.xlsx",majorbypos)
